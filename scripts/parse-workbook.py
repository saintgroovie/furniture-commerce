#!/usr/bin/env python3
"""
Woodright price workbook parser.

Reads "Розничный прайс 18.03.2026.xlsx" and produces normalized JSON
for further content mapping.  This is a data-preparation step only —
it does NOT modify backend, storefront, or seed data.

Usage:
    python3 scripts/parse-workbook.py [path-to-xlsx]

Output:
    data/raw/workbook/parsed-sheets.json   — all parsed rows
    data/raw/workbook/parse-summary.json   — per-sheet statistics
    data/raw/workbook/parse-warnings.json  — rows with parse issues
"""

import json
import os
import re
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_WORKBOOK = Path.home() / "Downloads" / "Розничный прайс 18.03.2026.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "data" / "raw" / "workbook"

# ---------------------------------------------------------------------------
# Collection normalization table
# ---------------------------------------------------------------------------
COLLECTION_NORM = {
    "ОЛИВЕР - ЧЕРНЫЙ": "oliver",
    "ГРИНВИЧ": "greenwich",
    "ВВ": "willie-winkie",
    "ОКСФОРД": "oxford",
    "ПРОВАНС": "provence",
    "ПРИНЦЕССА РОЗА": "princess-rose",
    "КАНТРИ-ЛОНДОН-ПАРИЖ": "country-london-paris",
    "МОНЧЕЛСИ": "monchelsea",
    "Аксессуары": "accessories",
    "Детали": "details",
    "Спецзаказ": "special-order",
}

COLLECTION_PREFIX = {
    "ОЛИВЕР - ЧЕРНЫЙ": "OL",
    "ГРИНВИЧ": "GR",
    "ВВ": "WW",
    "ОКСФОРД": "OX",
    "ПРОВАНС": "PV",
    "ПРИНЦЕССА РОЗА": "PR",
    "КАНТРИ-ЛОНДОН-ПАРИЖ": "CO",
    "МОНЧЕЛСИ": "MN",
    "Аксессуары": "A",
    "Детали": None,
    "Спецзаказ": "S",
}

# ---------------------------------------------------------------------------
# Per-sheet column mapping
# Each entry: (code_col, name_col, qty_col, dims_col, data_start_row,
#              price_columns: list of (col, label, price_type))
# Columns are 1-indexed to match openpyxl.
# ---------------------------------------------------------------------------
SHEET_CONFIG = {
    "ОЛИВЕР - ЧЕРНЫЙ": {
        "code_col": 3,
        "name_col": 4,
        "qty_col": 5,
        "dims_col": 9,
        "data_start": 7,
        "prices": [
            (6, "Массив", "base"),
            (7, "Черный BI (+15%)", "tier"),
            (8, "ЛДСП", "base"),
        ],
    },
    "ГРИНВИЧ": {
        "code_col": 2,
        "name_col": 3,
        "qty_col": 4,
        "dims_col": 7,
        "data_start": 6,
        "prices": [
            (5, "Массив", "base"),
            (6, "ЛДСП", "base"),
        ],
    },
    "ВВ": {
        "code_col": 3,
        "name_col": 4,
        "qty_col": 5,
        "dims_col": 10,
        "data_start": 11,
        "prices": [
            (6, "Роспись Tier 1", "tier"),
            (7, "Роспись Tier 2", "tier"),
            (8, "Роспись Tier 3", "tier"),
            (9, "ЛДСП", "base"),
        ],
    },
    "ОКСФОРД": {
        "code_col": 2,
        "name_col": 3,
        "qty_col": 4,
        "dims_col": 6,
        "data_start": 5,
        "prices": [
            (5, "Единая цена", "base"),
        ],
    },
    "ПРОВАНС": {
        "code_col": 3,
        "name_col": 4,
        "qty_col": 5,
        "dims_col": 8,
        "data_start": 6,
        "prices": [
            (6, "Прованс белый (ПБ)", "base"),
            (7, "Прованс тёмный (ПТ)", "base"),
        ],
    },
    "ПРИНЦЕССА РОЗА": {
        "code_col": 2,
        "name_col": 3,
        "qty_col": 4,
        "dims_col": 6,
        "data_start": 6,
        "prices": [
            (5, "Единая цена", "base"),
        ],
    },
    "КАНТРИ-ЛОНДОН-ПАРИЖ": {
        "code_col": 3,
        "name_col": 4,
        "qty_col": 5,
        "dims_col": 9,
        "data_start": 5,
        "prices": [
            (6, "Кантри", "base"),
            (7, "Лондон", "base"),
            (8, "Париж", "base"),
        ],
    },
    "МОНЧЕЛСИ": {
        "code_col": 2,
        "name_col": 4,
        "qty_col": 5,
        "dims_col": 8,
        "data_start": 11,
        "prices": [
            (6, "Массив", "base"),
            (7, "ЛДСП", "base"),
        ],
    },
    "Аксессуары": {
        "code_col": 2,
        "name_col": 3,
        "qty_col": 4,
        "dims_col": 6,
        "data_start": 5,
        "prices": [
            (5, "Единая цена", "base"),
        ],
    },
    "Детали": {
        "code_col": 2,
        "name_col": 3,
        "qty_col": 4,
        "dims_col": 6,
        "data_start": 5,
        "prices": [
            (5, "Единая цена", "base"),
        ],
    },
    "Спецзаказ": {
        "code_col": 2,
        "name_col": 3,
        "qty_col": 4,
        "dims_col": 6,
        "data_start": 5,
        "prices": [
            (5, "Цена/Коэффициент", "unknown"),
        ],
    },
}

# ---------------------------------------------------------------------------
# Category inference from product name
# ---------------------------------------------------------------------------
CATEGORY_PATTERNS = [
    (r"(?i)комплекс|замок", "complex", "Комплексы"),
    (r"(?i)банкетк[аи]", "bench", "Банкетки"),
    (r"(?i)диван", "sofa", "Диваны"),
    (r"(?i)комод", "dresser", "Комоды"),
    (r"(?i)кроват(?:ь|ка)|трансформер", "bed", "Кровати"),
    (r"(?i)бортик", "bed-guard", "Бортики"),
    (r"(?i)зеркал[оа]", "mirror", "Зеркала"),
    (r"(?i)полк[аи]|полочк", "shelf", "Полки"),
    (r"(?i)этажерк[аи]", "bookcase", "Стеллажи"),
    (r"(?i)стеллаж", "bookcase", "Стеллажи"),
    (r"(?i)стол(?:\s|$|ик|ешниц|-)", "table", "Столы"),
    (r"(?i)[сc]тол\s+письм", "table", "Столы"),
    (r"(?i)стул|табурет", "chair", "Стулья"),
    (r"(?i)тумб(?:а|ы|очк)", "nightstand", "Тумбы"),
    (r"(?i)шкаф|гардероб|витрин", "wardrobe", "Шкафы"),
    (r"(?i)секция\s+угловая", "wardrobe", "Шкафы"),
    (r"(?i)консоль", "console", "Консоли"),
    (r"(?i)надстройк", "desk-hutch", "Надстройки"),
    (r"(?i)сундук", "chest", "Сундуки"),
    (r"(?i)шкатулк", "jewelry-box", "Шкатулки"),
    (r"(?i)подушк", "cushion", "Подушки"),
    (r"(?i)валик", "bolster", "Валики"),
    (r"(?i)часы", "clock", "Часы"),
    (r"(?i)кресл", "armchair", "Кресла"),
    (r"(?i)обувниц", "shoe-rack", "Обувницы"),
    (r"(?i)вешалк", "hanger", "Вешалки"),
    (r"(?i)каркас.*балдахин", "canopy-frame", "Каркасы"),
    (r"(?i)(?:стеновая\s+)?панел[ьи]", "wall-panel", "Панели"),
    (r"(?i)ручк[аи]|ручка-", "handle", "Ручки"),
    (r"(?i)штанг[аи]", "rod", "Штанги"),
    (r"(?i)защ[ёе]лк", "latch", "Фурнитура"),
    (r"(?i)замок|замк", "lock", "Замки"),
    (r"(?i)филенк", "panel", "Филёнки"),
    (r"(?i)двер[ьи]", "door", "Двери"),
    (r"(?i)рамк[аи]", "frame", "Рамки"),
    (r"(?i)ящик", "toy-box", "Ящики"),
    (r"(?i)чехол", "cover", "Чехлы"),
    (r"(?i)лестниц", "ladder", "Лестницы"),
    (r"(?i)матрас", "mattress", "Матрасы"),
    (r"(?i)перекрас|обивк|отказ от росписи|увеличен", "modification", "Модификации"),
]


def infer_category(name):
    # type: (str) -> Tuple[Optional[str], Optional[str]]
    """Return (category_normalized, category_raw) from product name."""
    if not name:
        return None, None
    for pattern, slug, label in CATEGORY_PATTERNS:
        if re.search(pattern, name):
            return slug, label
    return None, None


# ---------------------------------------------------------------------------
# Dimension parsing
# ---------------------------------------------------------------------------
SEP = r"[\s*хxХX×_]+"  # any dimension separator

DIM_PATTERNS = [
    # В.1000 х Ш.650 х Гл.30 — with explicit В/Ш/Гл|Дл labels, any separator
    # Handles: х, *, space, mixed separators
    # Handles: slash ranges like В.750/1150 — captures first number
    re.compile(
        r"[ВBв][.\s]*(\d+)(?:/\d+)?\s*" + SEP +
        r"[Шш][.\s]*(\d+)(?:/\d+)?\s*" + SEP +
        r"(?:Гл|ГЛ|гл|г|Дл|ДЛ|дл)[.\s]*(\d+)"
    ),
    # В1529_Ш1862_Гл2146 — underscore/space separator, no dots
    re.compile(
        r"[ВBв](\d+)(?:/\d+)?[_\s]*[Шш](\d+)(?:/\d+)?[_\s]*(?:Гл|ГЛ|гл|г|Дл|ДЛ|дл)(\d+)"
    ),
    # Ш.398хГл.540хВ.18 — non-standard order W×D×H
    re.compile(
        r"[Шш][.\s]*(\d+)\s*" + SEP +
        r"(?:Гл|гл|г)[.\s]*(\d+)\s*" + SEP +
        r"[ВBв][.\s]*(\d+)"
    ),
    # дл.1084*гл 544*в18 — Дл×Гл×В lowercase
    re.compile(
        r"(?:дл|Дл)[.\s]*(\d+)\s*" + SEP +
        r"(?:гл|Гл)[.\s]*(\d+)\s*" + SEP +
        r"[вВ][.\s]*(\d+)"
    ),
    # Дл.152 х Ш.22 х Гл.32 — Дл×Ш×Гл format
    re.compile(
        r"(?:Дл|дл|ДЛ)[.\s]*(\d+)\s*" + SEP +
        r"[Шш][.\s]*(\d+)\s*" + SEP +
        r"(?:Гл|гл|г)[.\s]*(\d+)"
    ),
    # В.1670.Ш.650.Гл.391 — dots as separators (no space/х between groups)
    re.compile(
        r"[ВBв][.\s]*(\d+)(?:/\d+)?[.\s]*[Шш][ю.]?\s*(\d+)[.\s]*(?:Гл|гл|г|дл|Дл)[.\s]*(\d+)"
    ),
    # В. 1150 x 1482 x 1990 (no Ш/Гл labels, just three numbers after В)
    re.compile(
        r"[ВBв][.\s]*(\d+)(?:/\d+)?\s*" + SEP + r"(\d+)\s*" + SEP + r"(\d+)"
    ),
    # 500*395*440 or 900*1850*H200 — plain 3-number pattern
    re.compile(r"(\d{2,5})\s*[*×xХ]\s*(\d{2,5})\s*[*×xХ]\s*[HВвh]?(\d{2,5})"),
]

# Patterns that output groups in non-standard order need remapping.
# Key = pattern index, Value = lambda (g1, g2, g3) → (H, W, D)
_REMAP = {
    2: lambda v: (v[2], v[0], v[1]),   # Ш×Гл×В → (W,D,H) → remap to (H,W,D)
    3: lambda v: (v[2], v[1], v[0]),   # дл×гл×в → (L,D,H) → remap to (H,L,D)
    4: lambda v: (v[0], v[1], v[2]),   # Дл×Ш×Гл → treat Дл as H for consistency
}


def parse_dimensions(raw):
    # type: (Optional[str]) -> Optional[Dict]
    """Try to extract {height, width, depth} in mm from raw string."""
    if not raw or not isinstance(raw, str):
        return None
    text = str(raw).strip()
    if not any(c.isdigit() for c in text):
        return None
    for i, pat in enumerate(DIM_PATTERNS):
        m = pat.search(text)
        if m:
            vals = [int(x) for x in m.groups()]
            if i in _REMAP:
                h, w, d = _REMAP[i](vals)
                return {"height_mm": h, "width_mm": w, "depth_mm": d}
            return {"height_mm": vals[0], "width_mm": vals[1], "depth_mm": vals[2]}
    return None



# ---------------------------------------------------------------------------
# Price classification for Спецзаказ
# ---------------------------------------------------------------------------
MULTIPLIER_THRESHOLD = 10


def classify_special_price(val) -> str:
    """Determine if a price value is multiplier, absolute, or unknown."""
    if val is None:
        return "unknown"
    try:
        v = float(val)
    except (TypeError, ValueError):
        return "unknown"
    if v == 0:
        return "zero"
    if v < MULTIPLIER_THRESHOLD:
        return "modifier"
    return "absolute"


# ---------------------------------------------------------------------------
# Article normalization
# ---------------------------------------------------------------------------
SKIP_CODES = {"КОД", "АРТИКУЛ", "ТОН", "НАИМЕНОВАНИЕ", "КОЛ-ВО", "ЦЕНА"}


def is_header_value(val: str) -> bool:
    return val.upper().strip() in SKIP_CODES or "наименование" in val.lower()


def normalize_article(raw_code, sheet_name):
    # type: (str, str) -> Optional[str]
    """Normalize article code to {PREFIX}-{numeric} form."""
    if not raw_code or not isinstance(raw_code, str):
        return None
    code = raw_code.strip()
    if is_header_value(code):
        return None

    prefix = COLLECTION_PREFIX.get(sheet_name)

    # Детали and Спецзаказ already have their own prefixes
    if sheet_name in ("Детали", "Спецзаказ"):
        return code

    # If code already has a recognizable prefix, keep it
    if re.match(r"[A-Za-z]", code):
        return code

    # Bare numeric code — prepend collection prefix
    if prefix and re.match(r"\d", code):
        return f"{prefix}-{code}"

    return code


# ---------------------------------------------------------------------------
# Core row parser
# ---------------------------------------------------------------------------
def cell_val(ws, row: int, col: int):
    """Read cell value, handling merged cells gracefully."""
    return ws.cell(row=row, column=col).value


def parse_sheet(wb, sheet_name):
    # type: (...) -> List[Dict]
    """Parse one worksheet according to its config. Returns list of row dicts."""
    cfg = SHEET_CONFIG.get(sheet_name)
    if cfg is None:
        return []

    ws = wb[sheet_name]
    rows = []

    is_accessory = sheet_name == "Аксессуары"
    is_detail = sheet_name == "Детали"
    is_special = sheet_name == "Спецзаказ"

    for r in range(cfg["data_start"], ws.max_row + 1):
        code_raw = cell_val(ws, r, cfg["code_col"])
        name_raw = cell_val(ws, r, cfg["name_col"])

        # Skip empty / header rows
        if not name_raw:
            continue
        name_raw = str(name_raw).strip()
        if not name_raw:
            continue

        code_str = str(code_raw).strip() if code_raw else None
        if code_str and is_header_value(code_str):
            continue

        # At least one price must be present, or it's a note row
        has_any_price = False
        for pcol, _, _ in cfg["prices"]:
            v = cell_val(ws, r, pcol)
            if v is not None and v != "" and v != 0:
                has_any_price = True
                break

        # For Детали some rows lack price but are valid (handles)
        if not has_any_price and not is_detail:
            # Could be a note row — check if code looks valid
            if not code_str:
                continue

        warnings = []  # type: List[str]

        # Article
        code_normalized = normalize_article(code_str, sheet_name) if code_str else None
        if not code_str and name_raw:
            warnings.append("missing_article")
        if code_str and not code_normalized:
            warnings.append("article_normalization_failed")

        # Collection
        coll_normalized = COLLECTION_NORM.get(sheet_name, sheet_name.lower())

        # Dimensions
        dims_raw_val = cell_val(ws, r, cfg["dims_col"])
        dims_raw = str(dims_raw_val).strip() if dims_raw_val else None
        dims_normalized = parse_dimensions(dims_raw)
        if dims_raw and not dims_normalized:
            warnings.append("dimensions_parse_failed")

        # Category from name
        cat_norm, cat_raw = infer_category(name_raw)
        if not cat_norm:
            warnings.append("category_inference_failed")

        # Prices
        price_entries = []
        for pcol, plabel, ptype in cfg["prices"]:
            pval = cell_val(ws, r, pcol)

            # Determine effective price type
            effective_type = ptype
            if is_special:
                effective_type = classify_special_price(pval)

            price_raw = pval
            price_normalized = None
            if pval is not None:
                try:
                    price_normalized = float(pval)
                except (TypeError, ValueError):
                    warnings.append(f"price_parse_failed:{plabel}")

            price_entries.append({
                "label": plabel,
                "price_raw": price_raw,
                "price_normalized": price_normalized,
                "price_type": effective_type,
            })

        # Lead price for the record (first non-zero base/tier price)
        lead_price_raw = None
        lead_price_normalized = None
        lead_price_type = "unknown"
        for pe in price_entries:
            if pe["price_normalized"] and pe["price_normalized"] > 0:
                lead_price_raw = pe["price_raw"]
                lead_price_normalized = pe["price_normalized"]
                lead_price_type = pe["price_type"]
                break

        # Спецзаказ: also capture lead time from column 6 (same as dims_col)
        notes_raw = None
        if is_special:
            lead_time = cell_val(ws, r, 6)
            if lead_time and isinstance(lead_time, str) and "дн" in lead_time.lower():
                notes_raw = lead_time.strip()

        is_ambiguous = bool(warnings) or (is_special and lead_price_type == "unknown")

        row_data = {
            "source_sheet": sheet_name,
            "row_index": r,
            "collection_name_raw": sheet_name,
            "collection_name_normalized": coll_normalized,
            "product_code_raw": code_str,
            "product_code_normalized": code_normalized,
            "product_name_raw": name_raw,
            "product_name_canonical": name_raw,
            "category_raw": cat_raw,
            "category_normalized": cat_norm,
            "dimensions_raw": dims_raw,
            "dimensions_normalized": dims_normalized,
            "price_raw": lead_price_raw,
            "price_normalized": lead_price_normalized,
            "price_type": lead_price_type,
            "price_variants": price_entries,
            "notes_raw": notes_raw,
            "is_accessory": is_accessory,
            "is_detail": is_detail,
            "is_special_order": is_special,
            "is_ambiguous": is_ambiguous,
            "parse_warnings": warnings,
        }
        rows.append(row_data)

    return rows


# ---------------------------------------------------------------------------
# Summary builder
# ---------------------------------------------------------------------------
def build_summary(all_rows):
    # type: (List[Dict]) -> Dict
    sheets = {}
    total = 0
    total_warnings = 0
    total_ambiguous = 0

    for row in all_rows:
        sn = row["source_sheet"]
        if sn not in sheets:
            sheets[sn] = {
                "sheet_name": sn,
                "collection_normalized": row["collection_name_normalized"],
                "total_rows": 0,
                "rows_with_warnings": 0,
                "ambiguous_rows": 0,
                "rows_without_article": 0,
                "rows_without_price": 0,
                "unique_categories": set(),
                "price_labels": set(),
            }
        s = sheets[sn]
        s["total_rows"] += 1
        total += 1
        if row["parse_warnings"]:
            s["rows_with_warnings"] += 1
            total_warnings += 1
        if row["is_ambiguous"]:
            s["ambiguous_rows"] += 1
            total_ambiguous += 1
        if not row["product_code_raw"]:
            s["rows_without_article"] += 1
        if not row["price_normalized"]:
            s["rows_without_price"] += 1
        if row["category_normalized"]:
            s["unique_categories"].add(row["category_normalized"])
        for pv in row.get("price_variants", []):
            s["price_labels"].add(pv["label"])

    # Convert sets to sorted lists for JSON
    summary_sheets = []
    for sn in SHEET_CONFIG:
        if sn in sheets:
            s = sheets[sn]
            s["unique_categories"] = sorted(s["unique_categories"])
            s["price_labels"] = sorted(s["price_labels"])
            summary_sheets.append(s)

    return {
        "workbook_file": "Розничный прайс 18.03.2026.xlsx",
        "total_rows_parsed": total,
        "total_rows_with_warnings": total_warnings,
        "total_ambiguous_rows": total_ambiguous,
        "sheets": summary_sheets,
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    workbook_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_WORKBOOK
    if not workbook_path.exists():
        print(f"ERROR: Workbook not found at {workbook_path}", file=sys.stderr)
        sys.exit(1)

    print(f"Reading workbook: {workbook_path}")
    wb = openpyxl.load_workbook(str(workbook_path), data_only=True)

    all_rows = []  # type: List[Dict]
    for sheet_name in wb.sheetnames:
        if sheet_name not in SHEET_CONFIG:
            print(f"  SKIP: '{sheet_name}' (no config)")
            continue
        rows = parse_sheet(wb, sheet_name)
        all_rows.extend(rows)
        print(f"  OK: '{sheet_name}' → {len(rows)} rows")

    # Build summary
    summary = build_summary(all_rows)
    print(f"\nTotal: {summary['total_rows_parsed']} rows parsed, "
          f"{summary['total_rows_with_warnings']} with warnings, "
          f"{summary['total_ambiguous_rows']} ambiguous")

    # Collect warnings
    warning_rows = [r for r in all_rows if r["parse_warnings"]]

    # Ensure output dir
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Write parsed-sheets.json
    out_parsed = OUTPUT_DIR / "parsed-sheets.json"
    with open(out_parsed, "w", encoding="utf-8") as f:
        json.dump(all_rows, f, ensure_ascii=False, indent=2, default=str)
    print(f"  Wrote: {out_parsed} ({len(all_rows)} rows)")

    # Write parse-summary.json
    out_summary = OUTPUT_DIR / "parse-summary.json"
    with open(out_summary, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
    print(f"  Wrote: {out_summary}")

    # Write parse-warnings.json
    out_warnings = OUTPUT_DIR / "parse-warnings.json"
    with open(out_warnings, "w", encoding="utf-8") as f:
        json.dump(warning_rows, f, ensure_ascii=False, indent=2, default=str)
    print(f"  Wrote: {out_warnings} ({len(warning_rows)} rows)")

    print("\nDone.")


if __name__ == "__main__":
    main()
